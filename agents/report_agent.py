"""
report_agent.py  ─  ReportAgent 노드 (registry 버전)
개인화 대체 메뉴(personal_menus) 시트 추가
"""

import os
import pandas as pd
import registry
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from state import MealPlanState

TYPE_COLORS = {
    "HK형": "FCE4D6", "DHK형": "FCE4D6", "K형": "FCE4D6", "DK형": "FCE4D6",
    "DH형": "FFF2CC", "D형":   "FFF2CC",
    "H형":  "EBF1DE", "일반형": "FFFFFF",
}

def get_type_color(disease_label: str) -> str:
    if disease_label in TYPE_COLORS:
        return TYPE_COLORS[disease_label]
    if "M" in disease_label:
        return "E1D5E7"
    if "K" in disease_label:
        return "FCE4D6"
    if "D" in disease_label and "H" in disease_label:
        return "FFF2CC"
    if "D" in disease_label:
        return "FFF2CC"
    if "H" in disease_label:
        return "EBF1DE"
    return "FFFFFF"

SLOT_CATS = [
    ("밥","밥"),("국","국"),("주찬","주찬"),
    ("부찬1","부찬"),("부찬2","부찬"),("김치","김치"),
]


def report_agent(state: MealPlanState) -> dict:
    print("\n[ReportAgent] 보고서 생성 시작...")

    if not state.get("df_menu_records"):
        print("  [경고] df_menu_records 없음 — 건너뜀")
        return {"report_paths": {}, "messages": ["[ReportAgent] df_menu 없음"]}

    df = pd.DataFrame(
        state["df_menu_records"],
        columns=state["df_menu_columns"]
    )

    patients = registry.get(state["patients_key"]) if state.get("patients_key") else []

    paths: dict = {}

    meal_path = _save_meal_plan_excel(df, state.get("recommend_map") or {})
    paths["meal_plan"] = meal_path

    serving_path = _save_serving_excel(
        df, patients,
        state.get("serving_map")    or {},
        state.get("constraint_key"),
        state.get("personal_menus") or {},
        state.get("personalize_reasons") or {},
    )
    paths["serving"] = serving_path

    openai_key = os.getenv("OPENAI_API_KEY", "")
    if openai_key:
        guide_path = _save_cooking_guide(df, patients, openai_key)
        paths["cooking"] = guide_path
    else:
        print("  OPENAI_API_KEY 미설정 → 조리 지침서 건너뜀")

    print(f"[ReportAgent] 완료 — {list(paths.keys())}")
    return {
        "report_paths": paths,
        "messages":     [f"[ReportAgent] 보고서 생성 완료: {list(paths.values())}"],
    }


def _hdr(ws, addr, val, bg="1F497D", fg="FFFFFF", size=10):
    c = ws[addr]
    c.value = val
    c.font  = Font(name="맑은 고딕", bold=True, color=fg, size=size)
    c.fill  = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _wrt(ws, addr, val, bg=None, bold=False, size=9, color="000000", align="center"):
    c = ws[addr]
    c.value = val
    c.font  = Font(name="맑은 고딕", bold=bold, size=size, color=color)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def _border(ws, r1, r2, c1, c2):
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _save_meal_plan_excel(df: pd.DataFrame, recommend_map: dict,
                          path="식단표_28일.xlsx") -> str:
    """
    [수정 — 2026-07-01]
    - 영양소(열량/나트륨/단백질) 컬럼이 비어 보이는 문제 방어: None/누락 값을
      0으로 채워서 표시.
    - 셀 배경색: "권장재료포함수"에 따라 초록색 계열로 강조하던 것을
      제거하고 흰색으로 통일.
    - "권장재료포함메뉴"/"권장재료포함수" 컬럼명을 "포함된 식재료"/
      "식재료 개수"로 변경(표시명만 변경 — 실제 값이 "메뉴 개수"가 아니라
      "식재료 개수" 의미를 갖게 하려면 이 값을 산출하는 상위 로직
      (meal_plan_agent.py 등)도 같이 확인이 필요함. 이 파일은 표시만 담당).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "28일_식단표"
    ws.sheet_view.showGridLines = False

    cols   = ["일차","끼니","밥","국","주찬","부찬1","부찬2","김치",
              "열량(kcal)","나트륨(mg)","단백질(g)","비용(원)",
              "포함된 식재료","식재료 개수"]
    widths = [6,6,12,14,14,14,14,10,10,10,9,9,40,10]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    _hdr(ws, "A1", "28일 최적 식단표", size=12)
    ws.row_dimensions[1].height = 28

    for i, h in enumerate(cols, 1):
        _hdr(ws, f"{get_column_letter(i)}2", h, bg="2E5A9C", size=9)
    ws.row_dimensions[2].height = 20

    display_source = {
        "포함된 식재료": "권장재료포함메뉴",
        "식재료 개수":   "권장재료포함수",
    }

    if "권장재료포함수" not in df.columns:
        df = df.copy()
        df["권장재료포함메뉴"] = "-"
        df["권장재료포함수"]   = 0

    for r_idx, (_, row) in enumerate(df.iterrows(), 3):
        # [수정 — 2026-07-01] df.itertuples()는 "열량(kcal)" 같이 괄호가
        # 포함된 컬럼명을 파이썬 식별자로 쓸 수 없어 _3/_4/_5... 같은
        # 위치 기반 이름으로 자동 치환함(pandas 네임드튜플 제약). 그 결과
        # d.get("열량(kcal)", "")가 항상 기본값("")을 반환하고, 그게 바로
        # 아래 "빈 값이면 0 표시" 로직에 걸려 열량/나트륨/단백질/비용
        # 4개 컬럼이 전부 0으로 찍히던 버그였음(Neo4j 데이터 문제가
        # 아니었음). iterrows()는 Series를 반환해 컬럼명을 그대로 키로
        # 쓸 수 있어 이 문제가 없음.
        d = row.to_dict()

        for i, col in enumerate(cols, 1):
            source_col = display_source.get(col, col)
            val = d.get(source_col, "")

            if col in ("열량(kcal)", "나트륨(mg)", "단백질(g)", "비용(원)", "식재료 개수"):
                if val is None or val == "":
                    val = 0

            _wrt(ws, f"{get_column_letter(i)}{r_idx}", val, bg=None,
                 align="left" if col == "포함된 식재료" else "center")
        ws.row_dimensions[r_idx].height = 16

    _border(ws, 2, len(df)+2, 1, len(cols))
    wb.save(path)
    print(f"  식단표_28일.xlsx 저장 완료")
    return path


def _save_serving_excel(df: pd.DataFrame, patients: list,
                        serving_map: dict, constraint_key: str,
                        personal_menus: dict,
                        personalize_reasons: dict = None,
                        path="개인별_배식량.xlsx") -> str:
    personalize_reasons = personalize_reasons or {}
    constraint = registry.get(constraint_key) if constraint_key and registry.has(constraint_key) else None

    wb  = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "개인별_배식량"
    ws1.sheet_view.showGridLines = False

    col_hdrs = ["이름","질환유형","일차","끼니","ratio",
                "밥(g)","국(ml)","주찬(g)","부찬1(g)","부찬2(g)","김치(g)",
                "예상열량","예상단백질","예상나트륨","예상탄수화물",
                "열량OK","단백질OK","나트륨OK"]
    widths   = [10,8,6,6,6,7,7,7,7,7,7,8,8,8,9,6,6,6]

    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.merge_cells(f"A1:{get_column_letter(len(col_hdrs))}1")
    _hdr(ws1, "A1", "개인별 배식량 및 예상 영양소", bg="1F497D", size=12)
    ws1.row_dimensions[1].height = 28
    for i, h in enumerate(col_hdrs, 1):
        _hdr(ws1, f"{get_column_letter(i)}2", h, bg="2E5A9C", size=9)

    # [추가 — 2026-07-01] 실제 배식 시 0.98 같은 소수점 ratio는 눈대중으로
    # 맞추기 어려움. 현실적으로 담아줄 수 있는 0.25 단위 눈금으로 반올림함.
    RATIO_STEPS = [0.25, 0.5, 0.75, 1.0, 1.25, 1.5, 2.0]

    def _snap_ratio(ratio: float) -> float:
        if not ratio:
            return 1.0
        return min(RATIO_STEPS, key=lambda s: abs(s - ratio))

    rows_detail = []
    for _, menu_row in df.iterrows():
        day  = menu_row["일차"]
        meal = menu_row["끼니"]
        for p in patients:
            key_str = f"{p.name}||{day}||{meal}"
            srv       = serving_map.get(key_str, {})
            raw_ratio = srv.get("ratio", 1.0) or 1.0

            # [수정 — 2026-07-01] ratio만 반올림하고 배식량(g)·예상 영양소를
            # 그대로 두면 표 안에서 숫자끼리 서로 안 맞게 됨(예: ratio=1.0인데
            # 밥(g)은 0.98 기준 값). 반올림한 비율만큼 나머지 값도 같이
            # 스케일링해 표 전체가 내적으로 일관되게 함.
            snapped_ratio = _snap_ratio(raw_ratio)
            scale = (snapped_ratio / raw_ratio) if raw_ratio else 1.0

            def _scaled(val, _scale=scale):
                return round((val or 0) * _scale, 1)

            rice_g   = _scaled(srv.get("밥", srv.get("죽", 0)))
            soup_ml  = _scaled(srv.get("국", 0))
            main_g   = _scaled(srv.get("주찬", 0))
            side1_g  = _scaled(srv.get("부찬1", 0))
            side2_g  = _scaled(srv.get("부찬2", 0))
            kimchi_g = _scaled(srv.get("김치", 0))
            energy   = _scaled(srv.get("예상열량", 0))
            protein  = _scaled(srv.get("예상단백질", 0))
            sodium   = _scaled(srv.get("예상나트륨", 0))
            carb     = _scaled(srv.get("예상탄수화물", 0))

            c     = getattr(p, "constraint", None) or constraint
            e_min = getattr(c, "energy_min", 0)    or 0
            e_max = getattr(c, "energy_max", 9999) or 9999
            p_max = getattr(c, "protein_max", 9999) or 9999
            s_max = getattr(c, "sodium_max", 9999)  or 9999

            # 스케일링된(=표에 실제로 찍히는) 값 기준으로 OK 여부 재판정
            ok_e = e_min <= energy <= e_max
            ok_p = protein <= p_max
            ok_s = sodium <= s_max

            rows_detail.append([
                p.name,
                getattr(p, "disease_type_label", "-"),
                day, meal,
                snapped_ratio,
                rice_g, soup_ml, main_g, side1_g, side2_g, kimchi_g,
                energy, protein, sodium, carb,
                "✅" if ok_e else "⚠️",
                "✅" if ok_p else "⚠️",
                "✅" if ok_s else "⚠️",
            ])

    for r_idx, row_vals in enumerate(rows_detail, 3):
        disease = row_vals[1]
        bg = get_type_color(disease)
        for i, val in enumerate(row_vals, 1):
            color = "974706" if str(val) == "⚠️" else "000000"
            _wrt(ws1, f"{get_column_letter(i)}{r_idx}", val,
                 bg=bg, size=9, color=color)
        ws1.row_dimensions[r_idx].height = 14

    _border(ws1, 2, len(rows_detail)+2, 1, len(col_hdrs))

    if personal_menus:
        ws2 = wb.create_sheet("개인화_부찬대체")
        ws2.sheet_view.showGridLines = False

        p_hdrs   = ["이름", "일차", "끼니", "구분", "기존 메뉴", "대체 메뉴",
                    "배식ratio", "사유"]
        p_widths = [12, 8, 8, 10, 18, 18, 9, 26]
        for i, w in enumerate(p_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        ws2.merge_cells(f"A1:{get_column_letter(len(p_hdrs))}1")
        _hdr(ws2, "A1", "개인화 부찬 대체 지침 (조리팀용)", bg="375623", size=12)
        ws2.row_dimensions[1].height = 28
        for i, h in enumerate(p_hdrs, 1):
            _hdr(ws2, f"{get_column_letter(i)}2", h, bg="4E7C2F", size=9)
        ws2.row_dimensions[2].height = 20

        df_idx = {
            (row["일차"], row["끼니"], slot): row.get(slot, "")
            for _, row in df.iterrows()
            for slot in ["밥", "국", "주찬", "부찬1", "부찬2", "김치"]
        }

        REASON_KR = {"disease": "질환 위반 보정", "preference": "선호도 기반 대체"}
        REASON_BG = {"disease": "FCE4D6", "preference": "DDEBF7"}

        p_rows = []
        for key, changes in personal_menus.items():
            name, day, meal = key.split("||")
            reasons_for_key = {r["slot"]: r for r in personalize_reasons.get(key, [])}

            for slot, alt_menu in changes.items():
                detail = reasons_for_key.get(slot)
                if detail:
                    orig_menu = detail.get("from") or df_idx.get((day, meal, slot), "-")
                    reason_code = detail.get("reason", "preference")
                    reason_label = REASON_KR.get(reason_code, "기타")
                    full_detail = detail.get("detail", reason_label)
                    ratio = detail.get("ratio")
                    ratio_str = f"{ratio:.2f}" if ratio else "-"
                else:
                    orig_menu = df_idx.get((day, meal, slot), "-")
                    reason_code = "preference"
                    full_detail = "기피메뉴 대체"
                    ratio_str = "-"

                p_rows.append([
                    name, day, meal, REASON_KR.get(reason_code, "기타"),
                    f"{slot}: {orig_menu}", alt_menu, ratio_str, full_detail,
                    reason_code,
                ])

        meal_order = {"아침": 0, "점심": 1, "저녁": 2}
        p_rows.sort(key=lambda x: (
            x[0],
            int(x[1].replace("일", "")),
            meal_order.get(x[2], 9),
        ))

        for r_idx, row_vals in enumerate(p_rows, 3):
            reason_code = row_vals.pop()
            bg = REASON_BG.get(reason_code, "FFFFFF")

            for i, val in enumerate(row_vals, 1):
                bold = (i == 6)
                color = "375623" if i == 6 else "000000"
                _wrt(ws2, f"{get_column_letter(i)}{r_idx}", val,
                     bg=bg, bold=bold, size=9, color=color,
                     align="left" if i in (5, 6, 8) else "center")
            ws2.row_dimensions[r_idx].height = 16

        _border(ws2, 2, len(p_rows)+2, 1, len(p_hdrs))

        n_disease = sum(
            1 for reasons in personalize_reasons.values()
            for r in reasons if r.get("reason") == "disease"
        )
        n_pref = sum(
            1 for reasons in personalize_reasons.values()
            for r in reasons if r.get("reason") == "preference"
        )
        print(f"  개인화_부찬대체 시트 저장 완료 ({len(p_rows)}건 — "
              f"질환위반 {n_disease}건 / 선호도 {n_pref}건)")
    else:
        print("  개인화 대체 없음 — 시트 생략")

    wb.save(path)
    print(f"  개인별_배식량.xlsx 저장 완료 ({len(rows_detail)}행)")
    return path


def _save_cooking_guide(df: pd.DataFrame, patients: list,
                        api_key: str, path="조리_지침서.txt") -> str:
    from openai import OpenAI

    row = df[(df["일차"] == "1일") & (df["끼니"] == "점심")]
    if row.empty:
        return ""
    row = row.iloc[0]
    menu_summary = " / ".join([
        row["밥"], row["국"], row["주찬"],
        row["부찬1"], row["부찬2"], row["김치"]
    ])
    disease_labels = list({
        getattr(p, "disease_type_label", "일반형") for p in patients
    })

    prompt = f"""
노인요양시설 조리 지침서를 작성해 주세요.
[오늘 메뉴] {menu_summary}
[입소자 질환 유형] {', '.join(disease_labels)}

각 메뉴별로:
1. 조리 시 주의사항 (나트륨, 당, 식감 조절)
2. 질환별 배식 조정 포인트
3. 위생 및 온도 관리

간결하고 실용적으로 작성해 주세요.
"""
    client   = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=600, temperature=0.3,
    )
    guide_text = response.choices[0].message.content.strip()

    with open(path, "w", encoding="utf-8") as f:
        f.write(guide_text)
    print(f"  조리_지침서.txt 저장 완료")
    return path