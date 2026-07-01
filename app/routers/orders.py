"""
routers/orders.py — 발주(구매) 미리보기 + 엑셀 다운로드
================================================================================
GET /preview — JSON 형태로 발주 항목 미리보기 (프론트 src/pages/OrderExcel.tsx가 사용)
GET /export  — 거래명세서 스타일 엑셀 파일을 바로 스트리밍 다운로드

계산 로직 자체는 app/services/order_service.py에 있고, 이 라우터는
JSON으로 내보낼지 엑셀로 내보낼지만 다르게 처리함(같은 데이터, 다른 표현).

엑셀은 보고서 파일들(report_agent.py)과 달리 Supabase Storage에 올리지
않고 요청 시점에 즉시 생성해 바로 스트리밍함 — 저장해 둘 필요 없이 언제든
같은 run_id로 다시 요청하면 동일한 계산을 재현할 수 있는 파생 데이터이기
때문.

[수정 — 2026-07-01] _build_order_excel을 order_service.py의 새 스키마
(run_id, week_range, total_items, total_cost, items[menu_name/ingredient/
servings_used/total_weight_g/product_name/unit_price/estimated_cost])에
맞춰 다시 작성.
"""

import io
from datetime import datetime, timezone
from urllib.parse import quote

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from app.services.order_service import build_order_data

router = APIRouter()


@router.get("/preview")
def preview_order(run_id: str, week_offset: int = 0):
    try:
        return build_order_data(run_id, week_offset)
    except ValueError as e:
        raise HTTPException(404, str(e))
    except Exception as e:
        raise HTTPException(500, f"발주 계산 중 오류: {e}")


@router.get("/export")
def export_order_excel(run_id: str, week_offset: int = 0):
    try:
        data = build_order_data(run_id, week_offset)
    except ValueError as e:
        raise HTTPException(404, str(e))
    except Exception as e:
        raise HTTPException(500, f"발주 계산 중 오류: {e}")

    buf = _build_order_excel(data)

    filename = f"발주서_{data['week_range']}.xlsx"
    encoded_filename = quote(filename)
    content_disposition = (
        f"attachment; filename=\"order.xlsx\"; filename*=UTF-8''{encoded_filename}"
    )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition},
    )


def _build_order_excel(data: dict) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def hdr(ws, addr, val, bg="1F497D", fg="FFFFFF", size=10, bold=True):
        c = ws[addr]
        c.value = val
        c.font = Font(name="맑은 고딕", bold=bold, color=fg, size=size)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def cell(ws, addr, val, align="center", bold=False, size=9):
        c = ws[addr]
        c.value = val
        c.font = Font(name="맑은 고딕", bold=bold, size=size)
        c.alignment = Alignment(horizontal=align, vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래명세서"
    ws.sheet_view.showGridLines = False

    widths = [6, 16, 20, 12, 12, 20, 12, 14, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:I1")
    hdr(ws, "A1", "거 래 명 세 서", bg="1F497D", size=16)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A3:D3")
    cell(ws, "A3", f"대상 기간: {data['week_range']}", align="left", bold=True, size=11)
    ws.merge_cells("F3:I3")
    cell(ws, "F3", f"발행일: {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", align="left", size=10)

    ws.merge_cells("A4:D4")
    cell(ws, "A4", f"총 품목 수: {data['total_items']}건", align="left", size=10)
    ws.merge_cells("F4:I4")
    cell(ws, "F4", f"총 발주 예상 금액: {data['total_cost']:,}원", align="left", bold=True, size=10)

    table_hdrs = ["번호", "메뉴명", "재료", "사용 끼니수", "총 중량(g)",
                  "구매 상품(참고)", "단가(원)", "예상 비용(원)", "비고"]
    header_row = 6
    for i, h in enumerate(table_hdrs, 1):
        hdr(ws, f"{get_column_letter(i)}{header_row}", h, bg="2E5A9C", size=9)
    ws.row_dimensions[header_row].height = 20

    thin = Side(style="thin")
    row_idx = header_row + 1
    for i, item in enumerate(data["items"], 1):
        values = [
            i,
            item["menu_name"],
            item["ingredient"],
            item["servings_used"],
            item["total_weight_g"],
            item["product_name"] or "-",
            item["unit_price"] if item["unit_price"] is not None else "-",
            item["estimated_cost"] if item["estimated_cost"] is not None else "-",
            "",
        ]
        for col, val in enumerate(values, 1):
            align = "left" if col in (2, 3, 6) else "center"
            cell(ws, f"{get_column_letter(col)}{row_idx}", val, align=align)
        row_idx += 1

    # 합계 행
    ws.merge_cells(f"A{row_idx}:G{row_idx}")
    hdr(ws, f"A{row_idx}", "합계", bg="D9E1F2", fg="1F497D", size=10)
    cell(ws, f"H{row_idx}", data["total_cost"], bold=True, size=10)
    ws[f"H{row_idx}"].fill = PatternFill("solid", fgColor="D9E1F2")

    last_row = row_idx
    for r in ws.iter_rows(min_row=header_row, max_row=last_row, min_col=1, max_col=9):
        for c in r:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf